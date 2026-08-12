"""Phase 1 feasibility probe for Amazon Reviews'23.

This script measures three things and nothing else. It is a feasibility
measurement, not an analysis: it produces no estimates, writes no files, and
makes no modelling choices.

  1. What share of reviews carry a usable fit judgement in title+text, split
     into ran_small / true_to_size / ran_large? Labels come from the rule-based
     regex dictionary declared at the top of this file (DESIGN.md 5.1). A review
     matching more than one bucket is dropped as ambiguous and counted
     separately -- never guessed. Example matches are printed per bucket so that
     precision can be established by hand.

  2. What share of items expose a purchased size in the `details` field, and what
     do those size strings actually look like? DESIGN.md 5.3 warns they are
     unconstrained free text; this measures how bad that is, and how much of it
     a conservative normaliser can recover.

  3. What share of items yield a gender and a body half from the `categories`
     field under the DESIGN.md 1.3 assignment, and what are the four
     gender x half cell counts? The men's-lower cell is reported explicitly
     because DESIGN.md 5.5 makes it the binding constraint.

Data access is stream-only: nothing is written to disk and no full file is
fetched. The HuggingFace streaming reader is tried first; if the hub's dataset
script is unavailable, the probe falls back to reading the published .jsonl over
HTTPS line by line and stopping at the requested sample size. Both paths are
single-connection, sequential, and identified by User-Agent.

Usage:
    python amazon_fit_probe.py [--reviews 50000] [--items 30000]
                               [--category Amazon_Fashion]
"""

from __future__ import annotations

import argparse
import collections
import csv
import datetime
import gzip
import hashlib
import io
import json
import pathlib
import random
import re
import sys
import urllib.request

# ---------------------------------------------------------------------------
# Question 1: the fit dictionary (DESIGN.md 5.1)
#
# Rule-based and auditable, deliberately favouring precision over recall. Every
# pattern is applied to lowercased `title + " " + text`. A review is labelled
# only if exactly one bucket matches; two or more buckets means ambiguous, which
# is a drop, not a guess.
# ---------------------------------------------------------------------------

FIT_DICTIONARY: dict[str, list[str]] = {
    # Garment was smaller than the buyer expected -> fit_score -1
    "ran_small": [
        r"\bruns?\s+(?:a\s+)?(?:bit|little|tad|touch|lot|way|very|really|super|slightly|somewhat)?\s*small(?:er)?\b",
        r"\bran\s+(?:a\s+)?(?:bit|little|tad|touch|lot|way|very|really|super|slightly|somewhat)?\s*small(?:er)?\b",
        r"\bruns?\s+(?:about\s+)?(?:a\s+|one\s+|two\s+)?(?:full\s+)?sizes?\s+(?:too\s+)?small\b",
        # REMOVED 2026-08-08: r"\bsmall\s+(?:in\s+)?siz(?:e|ing)\b"
        # Symmetric counterpart of the ran_large removal below. "the small size
        # was more comfortable" names the size the buyer chose, not the fit the
        # garment delivered. Removed before the dictionary is frozen in
        # PREREGISTRATION.md, so this is a pre-registration edit, not a
        # post-hoc one. See docs/phase1-amazon-probe.md.
        r"\bsizing\s+runs?\s+small\b",
        r"\b(?:way|far|much|a\s+bit|a\s+little|too)\s+too\s+small\b",
        r"\btoo\s+small\b",
        r"\btoo\s+tight\b",
        r"\b(?:very|really|super|so|extremely)\s+tight\b",
        r"\b(?:a\s+)?(?:bit|little|tad|touch)\s+tight\b",
        r"\bsmaller\s+than\s+(?:i\s+)?(?:expected|anticipated|thought|usual|normal|advertised|described|pictured|the\s+size\s+chart)\b",
        r"\bsize\s+up\b",
        r"\bsizing\s+up\b",
        r"\bsized\s+up\b",
        r"\bgo\s+(?:a\s+|one\s+|two\s+)?sizes?\s+up\b",
        r"\bgo\s+up\s+(?:a\s+|one\s+|two\s+)?sizes?\b",
        r"\b(?:order|buy|get|purchase|choose|pick)\s+(?:a\s+|one\s+|two\s+|the\s+next\s+)?sizes?\s+(?:up|larger|bigger)\b",
        r"\b(?:order|buy|get|purchase|choose|pick)\s+(?:a\s+|one\s+|two\s+)?(?:larger|bigger)\s+sizes?\b",
        r"\bnext\s+size\s+up\b",
        r"\bneed(?:ed)?\s+(?:a\s+)?(?:larger|bigger)\s+size\b",
        r"\bwish\s+(?:i|I)?\s*(?:had\s+)?order(?:ed)?\s+(?:a\s+)?(?:larger|bigger)\b",
        r"\bcould\s?n[o']?t\s+(?:even\s+)?(?:get\s+it\s+)?(?:zip|button|pull)\b",
    ],
    # Fit was right -> fit_score 0
    "true_to_size": [
        r"\btrue\s+to\s+siz(?:e|ing)\b",
        r"\btrue-to-size\b",
        r"\btts\b",
        r"\bfits?\s+true\b",
        r"\bfit(?:s|ted)?\s+(?:me\s+)?(?:just\s+)?(?:perfect|perfectly)\b",
        r"\bperfect\s+fit\b",
        r"\bfits?\s+(?:me\s+)?(?:just\s+)?right\b",
        r"\bfit\s+(?:me\s+)?(?:just\s+)?right\b",
        r"\bfits?\s+as\s+expected\b",
        r"\bfits?\s+exactly\s+as\s+(?:expected|described|advertised)\b",
        r"\bsiz(?:e|ing)\s+(?:is|was|runs)\s+(?:spot\s+on|accurate|correct|right|as\s+expected)\b",
        r"\baccurate\s+siz(?:e|ing)\b",
        r"\bsizing\s+is\s+(?:on\s+point|spot\s+on)\b",
        r"\bno\s+need\s+to\s+size\s+(?:up|down)\b",
        r"\b(?:don'?t|do\s+not|no\s+need\s+to)\s+(?:need\s+to\s+)?(?:order|go|buy|get)\s+(?:a\s+)?size\s+(?:up|down)\b",
        r"\b(?:my|the)\s+(?:usual|normal|regular)\s+size\s+fit(?:s|ted)?\b",
        r"\bordered\s+my\s+(?:usual|normal|regular)\s+size\s+and\s+(?:it\s+)?fit\b",
    ],
    # Garment was larger than the buyer expected -> fit_score +1
    "ran_large": [
        r"\bruns?\s+(?:a\s+)?(?:bit|little|tad|touch|lot|way|very|really|super|slightly|somewhat)?\s*(?:large|larger|big|bigger)\b",
        r"\bran\s+(?:a\s+)?(?:bit|little|tad|touch|lot|way|very|really|super|slightly|somewhat)?\s*(?:large|larger|big|bigger)\b",
        r"\bruns?\s+(?:about\s+)?(?:a\s+|one\s+|two\s+)?(?:full\s+)?sizes?\s+(?:too\s+)?(?:large|big)\b",
        # REMOVED 2026-08-08: r"\b(?:large|big)\s+(?:in\s+)?siz(?:e|ing)\b"
        # Demonstrated false positive. In the CSJ hand-verification sample it
        # fired on "she can usually wear medium underwear, but finds the large
        # size to be more comfortable", where "the large size" is the size
        # purchased, not a statement that the garment ran large. It fired 120
        # times in the 50k Amazon_Fashion pass. Removed before the dictionary is
        # frozen in PREREGISTRATION.md. See docs/phase1-amazon-probe.md.
        r"\bsizing\s+runs?\s+(?:large|big)\b",
        r"\b(?:way|far|much|a\s+bit|a\s+little|too)\s+too\s+(?:big|large)\b",
        r"\btoo\s+(?:big|large)\b",
        r"\btoo\s+(?:loose|baggy|roomy|wide)\b",
        r"\b(?:very|really|super|so|extremely)\s+(?:baggy|loose)\b",
        r"\b(?:bigger|larger)\s+than\s+(?:i\s+)?(?:expected|anticipated|thought|usual|normal|advertised|described|pictured|the\s+size\s+chart)\b",
        r"\bsize\s+down\b",
        r"\bsizing\s+down\b",
        r"\bsized\s+down\b",
        r"\bgo\s+(?:a\s+|one\s+|two\s+)?sizes?\s+down\b",
        r"\bgo\s+down\s+(?:a\s+|one\s+|two\s+)?sizes?\b",
        r"\b(?:order|buy|get|purchase|choose|pick)\s+(?:a\s+|one\s+|two\s+|the\s+next\s+)?sizes?\s+(?:down|smaller)\b",
        r"\b(?:order|buy|get|purchase|choose|pick)\s+(?:a\s+|one\s+|two\s+)?smaller\s+sizes?\b",
        r"\bnext\s+size\s+down\b",
        r"\bneed(?:ed)?\s+(?:a\s+)?smaller\s+size\b",
        r"\bwish\s+(?:i|I)?\s*(?:had\s+)?order(?:ed)?\s+(?:a\s+)?smaller\b",
        r"\b(?:huge|enormous)\s+on\s+(?:me|him|her)\b",
        r"\bswimming\s+in\s+(?:it|them|this)\b",
        r"\blike\s+a\s+(?:tent|circus\s+tent)\b",
    ],
}

# Patterns that veto a bucket the include list would otherwise have matched.
# These exist because the include lists are surface-form matchers with no
# negation handling of their own.
NEGATION_PATTERNS: dict[str, list[str]] = {
    "ran_small": [
        r"\bnot\s+(?:at\s+all\s+)?too\s+(?:small|tight)\b",
        r"\bnot\s+(?:too\s+)?tight\s+at\s+all\b",
        r"\b(?:does\s?n[o']?t|do\s+not|did\s?n[o']?t|did\s+not)\s+run\s+small\b",
        r"\bno\s+need\s+to\s+size\s+up\b",
        r"\b(?:don'?t|do\s+not|did\s?n[o']?t)\s+need\s+to\s+size\s+up\b",
        r"\bno\s+need\s+to\s+(?:order|go|buy|get)\s+(?:a\s+)?size\s+up\b",
        r"\bnot\s+smaller\s+than\b",
    ],
    "true_to_size": [
        # The intervening-word forms are load-bearing: the Amazon_Fashion probe
        # of 2026-08-08 found "not made true to size" defeating a veto that
        # required "not" to sit directly before "true".
        r"\bnot\s+(?:\w+\s+){0,3}?true\s+to\s+siz(?:e|ing)\b",
        r"\b(?:is\s?n[o']?t|was\s?n[o']?t|are\s?n[o']?t|were\s?n[o']?t)\s+(?:\w+\s+){0,3}?true\s+to\s+siz(?:e|ing)\b",
        r"\bnot\s+(?:a\s+)?perfect\s+fit\b",
        r"\b(?:does\s?n[o']?t|did\s?n[o']?t|do\s+not|did\s+not)\s+fit\s+(?:me\s+)?(?:perfectly|right|as\s+expected)\b",
        r"\bnot\s+accurate\b",
        r"\bfar\s+from\s+true\s+to\s+size\b",
    ],
    "ran_large": [
        r"\bnot\s+(?:at\s+all\s+)?too\s+(?:big|large|loose|baggy)\b",
        r"\bnot\s+(?:too\s+)?(?:big|large|baggy|loose)\s+at\s+all\b",
        r"\b(?:does\s?n[o']?t|do\s+not|did\s?n[o']?t|did\s+not)\s+run\s+(?:big|large)\b",
        r"\bno\s+need\s+to\s+size\s+down\b",
        r"\b(?:don'?t|do\s+not|did\s?n[o']?t)\s+need\s+to\s+size\s+down\b",
        r"\bno\s+need\s+to\s+(?:order|go|buy|get)\s+(?:a\s+)?size\s+down\b",
        r"\bnot\s+(?:bigger|larger)\s+than\b",
    ],
}

# ---------------------------------------------------------------------------
# Question 2: purchased size in `details` (DESIGN.md 5.3)
# ---------------------------------------------------------------------------

# A details key is a size key if it is exactly "size" (strict) or mentions size
# without being a package/shipping dimension (loose). Both rates are reported.
SIZE_KEY_STRICT = re.compile(r"^size$")
SIZE_KEY_LOOSE = re.compile(r"\bsize\b")
SIZE_KEY_LOOSE_EXCLUDE = re.compile(
    r"package|shipping|chart|guide|one\s*size\s*fits|file|image"
)

# Conservative normalisation. Map only what is unambiguous; count and report
# everything else as dropped. Never guess (DESIGN.md 5.3).
ALPHA_SIZE_LADDER: dict[str, str] = {
    "xxxs": "XXXS", "3xs": "XXXS",
    "xxs": "XXS", "2xs": "XXS", "xx small": "XXS", "xx-small": "XXS",
    "xs": "XS", "x small": "XS", "x-small": "XS", "extra small": "XS",
    "s": "S", "sm": "S", "small": "S",
    "m": "M", "md": "M", "med": "M", "medium": "M",
    "l": "L", "lg": "L", "large": "L",
    "xl": "XL", "x large": "XL", "x-large": "XL", "extra large": "XL",
    "xxl": "XXL", "2xl": "XXL", "xx large": "XXL", "xx-large": "XXL",
    "xxxl": "XXXL", "3xl": "XXXL", "xxx large": "XXXL", "xxx-large": "XXXL",
    "xxxxl": "XXXXL", "4xl": "XXXXL",
    "xxxxxl": "XXXXXL", "5xl": "XXXXXL",
    "6xl": "XXXXXXL",
}

# "36W x 32L", "36x32", "34 W X 30 L"
WAIST_INSEAM = re.compile(
    r"^(\d{2})\s*w?\s*[x×*]\s*(\d{2})\s*l?$"
)
# "8", "US 8", "EU 38", "8.5"
PLAIN_NUMERIC = re.compile(r"^(?:us|eu|uk|jp)?\s*(\d{1,3}(?:\.\d)?)$")
# Recognised but deliberately not normalisable to a point on a ladder.
ONE_SIZE = re.compile(r"^one[\s-]*size(?:\s*fits\s*(?:all|most))?$")

# ---------------------------------------------------------------------------
# Question 3: gender and body half from `categories` (DESIGN.md 1.3)
# ---------------------------------------------------------------------------

# Storefront roots carry "Shoes & Jewelry" in the breadcrumb itself; matching
# garment or exclusion keywords against them would misclassify every item.
ROOT_SEGMENTS = {
    "clothing, shoes & jewelry",
    "clothing, shoes and jewelry",
    "clothing shoes & jewelry",
    "amazon fashion",
    "all beauty",
    "clothing",
    "novelty & more",
    "shops",
    "departments",
}

GENDER_PATTERNS: dict[str, str] = {
    "men": r"\bmen'?s?\b",
    "women": r"\bwomen'?s?\b|\bladies\b|\bjuniors\b",
}
CHILD_PATTERN = r"\bboys?'?\b|\bgirls?'?\b|\bbaby\b|\bbabies\b|\bkids?\b|\btoddler\b|\binfant\b|\bchildren'?s?\b"
UNISEX_PATTERN = r"\bunisex\b"

# Order matters: excluded is tested first, then lower, then upper.
EXCLUDED_HALF_PATTERN = (
    r"\bdress(?:es)?\b|\bjumpsuit|\bromper|\boverall|\bswim|\bbikini|\bcover[\s-]?up"
    r"|\bouterwear\b|\bjacket|\bcoat|\bparka|\bblazer|\bsuit\b|\bsuits\b"
    r"|\bunderwear\b|\bintimate|\blingerie\b|\bbra\b|\bbras\b|\bbralette|\bpanty|\bpanties"
    r"|\bboxer|\bbrief|\bthong|\bshapewear\b|\bsleepwear\b|\bpajama|\brobe"
    r"|\bsock|\bhosiery\b|\btight[s]?\b|\bstocking"
    r"|\bshoe|\bsneaker|\bboot|\bsandal|\bslipper|\bloafer|\bheel|\bfootwear\b"
    r"|\baccessor|\bjewelry\b|\bwatch|\bhat\b|\bhats\b|\bcap\b|\bcaps\b|\bbelt"
    r"|\bbag\b|\bbags\b|\bhandbag|\bwallet|\bscarf|\bscarve|\bglove|\bmitten"
    r"|\bsunglass|\beyewear\b|\btie\b|\bties\b|\bbowtie|\bcostume|\buniform"
)
LOWER_HALF_PATTERN = (
    r"\bjean\b|\bjeans\b|\bdenim\b|\bpant\b|\bpants\b|\btrouser|\bchino"
    r"|\bshort\b|\bshorts\b|\bskirt|\blegging|\bjogger|\bsweatpant|\bcargo\b"
)
UPPER_HALF_PATTERN = (
    r"\bt-?shirt|\btee\b|\btees\b|\bshirt|\bblouse|\bpolo|\bsweater|\bsweatshirt"
    r"|\bhoodie|\bhoodies\b|\bpullover|\bcardigan|\btop\b|\btops\b|\btank\b|\bcamisole|\bhenley"
)

# The primary three-step gradient of DESIGN.md 1.2. The three classes must be
# mutually exclusive or the ambiguity guard misfires: a bare r"\bshirt" also
# matches inside "t-shirts" (the hyphen is a word boundary), which made the leaf
# "t-shirts" look like it named two classes at once. So tee-forms are stripped
# out of a segment before the shirt rule is tested against the residue.
# "sweatshirt" and "undershirt" need no special casing -- r"\bshirt" cannot match
# them, as the preceding character is a word character.
GRADIENT_TEE = r"\bt-?shirts?\b|\btees?\b|\btanks?\b|\bcami(?:sole)?s?\b"
GRADIENT_TEE_STRIP = r"\bt-?shirts?\b|\btees?\b|\btanks?\b|\bcami(?:sole)?s?\b|\btops?\b"
GRADIENT_SHIRT = r"\bshirt|\bblouse|\bpolo\b|\bpolos\b|\bbutton-?(?:down|up)\b|\boxford\b"
GRADIENT_JEANS = r"\bjean\b|\bjeans\b|\bdenim\b|\bpant\b|\bpants\b|\btrouser|\bchino"

HF_REPO = "McAuley-Lab/Amazon-Reviews-2023"
HF_RESOLVE = f"https://huggingface.co/datasets/{HF_REPO}/resolve/main"
USER_AGENT = "the-vanity-gap-phase1-probe/1.0 (academic feasibility probe)"


# ---------------------------------------------------------------------------
# Compiled forms
# ---------------------------------------------------------------------------

FIT_RX = {
    bucket: [(p, re.compile(p)) for p in pats]
    for bucket, pats in FIT_DICTIONARY.items()
}
NEG_RX = {
    bucket: [re.compile(p) for p in pats]
    for bucket, pats in NEGATION_PATTERNS.items()
}
GENDER_RX = {g: re.compile(p) for g, p in GENDER_PATTERNS.items()}
CHILD_RX = re.compile(CHILD_PATTERN)
UNISEX_RX = re.compile(UNISEX_PATTERN)
EXCLUDED_RX = re.compile(EXCLUDED_HALF_PATTERN)
LOWER_RX = re.compile(LOWER_HALF_PATTERN)
UPPER_RX = re.compile(UPPER_HALF_PATTERN)
GRADIENT_TEE_RX = re.compile(GRADIENT_TEE)
GRADIENT_TEE_STRIP_RX = re.compile(GRADIENT_TEE_STRIP)
GRADIENT_SHIRT_RX = re.compile(GRADIENT_SHIRT)
GRADIENT_JEANS_RX = re.compile(GRADIENT_JEANS)

WS_RX = re.compile(r"\s+")


# ---------------------------------------------------------------------------
# Streaming readers -- nothing is written to disk, nothing is fetched in full
# ---------------------------------------------------------------------------

def _hf_url(config: str, category: str) -> str:
    if config.startswith("raw_review"):
        return f"{HF_RESOLVE}/raw/review_categories/{category}.jsonl"
    if config.startswith("raw_meta"):
        return f"{HF_RESOLVE}/raw/meta_categories/meta_{category}.jsonl"
    raise ValueError(f"unrecognised config: {config}")


def _iter_http_jsonl(url: str, limit: int):
    """Read a published .jsonl over HTTPS line by line and stop at `limit`."""
    request = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
    with urllib.request.urlopen(request) as response:
        stream = response
        if url.endswith(".gz") or response.headers.get("Content-Encoding") == "gzip":
            stream = gzip.GzipFile(fileobj=response)
        reader = io.TextIOWrapper(stream, encoding="utf-8", errors="replace")
        for i, line in enumerate(reader):
            if i >= limit:
                break
            line = line.strip()
            if line:
                yield json.loads(line)


def _file_size(url: str) -> int:
    request = urllib.request.Request(url, method="HEAD",
                                     headers={"User-Agent": USER_AGENT})
    with urllib.request.urlopen(request) as response:
        return int(response.headers["Content-Length"])


def _iter_range(url: str, start: int, limit: int):
    """Yield up to `limit` records beginning at byte offset `start`.

    The first line after a non-zero offset is a fragment of the record spanning
    the boundary, so it is discarded.
    """
    headers = {"User-Agent": USER_AGENT, "Range": f"bytes={start}-"}
    request = urllib.request.Request(url, headers=headers)
    with urllib.request.urlopen(request) as response:
        if start > 0 and response.status != 206:
            raise RuntimeError("server ignored Range; cannot block-sample")
        reader = io.TextIOWrapper(response, encoding="utf-8", errors="replace")
        if start > 0:
            reader.readline()
        yielded = 0
        for line in reader:
            if yielded >= limit:
                return
            line = line.strip()
            if not line:
                continue
            try:
                record = json.loads(line)
            except json.JSONDecodeError:
                continue
            yielded += 1
            yield record


def iter_records_spread(config: str, category: str, limit: int, blocks: int = 8):
    """Read `limit` records spread evenly across `blocks` disjoint file offsets.

    THE REASON THIS EXISTS: `iter_records` reads a PREFIX. The sampling-frame
    check of 2026-08-11 (`sampling_frame_probe.py`,
    docs/phase1-amazon-probe.md 6) established that these files are ORDERED --
    `verified_purchase` runs 64.88% -> 94.75% and mean review length 316 -> 142
    characters from head to tail, which is a time gradient. A prefix therefore
    samples the oldest reviews, not the corpus, and every rate derived from one
    is biased.

    Spreading the read across the file does not give a random sample either --
    within a block, records are still contiguous. It gives a systematic sample
    with a large number of well-separated start points, which removes the
    first-order bias from file order. State it that way; do not call it random.
    """
    url = _hf_url(config, category)
    size = _file_size(url)
    per_block = max(1, limit // blocks)
    print(f"  [reader] block-sampled: {blocks} blocks x {per_block:,} records "
          f"across {size / 1e9:.1f} GB")
    # The last block starts short of EOF so it can still yield a full block.
    span = size // blocks
    for index in range(blocks):
        start = min(index * span, max(0, size - span // 2))
        yielded = 0
        for record in _iter_range(url, start, per_block):
            yielded += 1
            yield record
        if yielded == 0 and index > 0:
            return


SPREAD_BLOCKS = 0   # 0 = prefix read (biased, see iter_records_spread); >0 = block-sampled


def iter_records(config: str, category: str, limit: int):
    """Yield up to `limit` records from a config, streaming only.

    Routes to `iter_records_spread` when SPREAD_BLOCKS is set, so every probe
    picks up block sampling without changing its own call sites.

    Tries the HuggingFace streaming reader first (DESIGN.md 3.1). If the hub's
    dataset script cannot be loaded -- recent `datasets` releases dropped script
    support -- falls back to reading the published .jsonl over HTTPS. Both paths
    are stream-only and stop early; neither downloads the file.
    """
    if SPREAD_BLOCKS:
        yield from iter_records_spread(config, category, limit, SPREAD_BLOCKS)
        return

    try:
        from datasets import load_dataset

        dataset = load_dataset(
            HF_REPO, config, split="full", streaming=True, trust_remote_code=True
        )
        print(f"  [reader] datasets.load_dataset streaming, config={config}")
        for i, record in enumerate(dataset):
            if i >= limit:
                break
            yield record
        return
    except Exception as exc:  # noqa: BLE001 -- fall back on any loader failure
        reason = f"{type(exc).__name__}: {exc}"[:180]
        print(f"  [reader] load_dataset unavailable ({reason})")

    url = _hf_url(config, category)
    print(f"  [reader] HTTPS line-stream fallback: {url}")
    yield from _iter_http_jsonl(url, limit)


# ---------------------------------------------------------------------------
# Question 1
# ---------------------------------------------------------------------------

def label_fit(title: str, text: str) -> tuple[str, dict[str, str]]:
    """Return (label, {bucket: matched_pattern}).

    label is one of the three buckets, "ambiguous" if two or more buckets fire,
    or "none". Ambiguous is a drop, never a guess (DESIGN.md 5.1).
    """
    blob = WS_RX.sub(" ", f"{title or ''} {text or ''}".lower()).strip()
    hits: dict[str, str] = {}
    for bucket, patterns in FIT_RX.items():
        matched = None
        for raw, rx in patterns:
            if rx.search(blob):
                matched = raw
                break
        if matched is None:
            continue
        if any(neg.search(blob) for neg in NEG_RX[bucket]):
            continue
        hits[bucket] = matched
    if len(hits) == 1:
        return next(iter(hits)), hits
    if len(hits) > 1:
        return "ambiguous", hits
    return "none", hits


def snippet(title: str, text: str, pattern: str, width: int = 110) -> str:
    blob = WS_RX.sub(" ", f"{title or ''} || {text or ''}").strip()
    match = re.search(pattern, blob.lower())
    if match is None:
        return blob[:width]
    start = max(0, match.start() - width // 2)
    end = min(len(blob), match.end() + width // 2)
    return ("..." if start else "") + blob[start:end] + ("..." if end < len(blob) else "")


def probe_reviews(category: str, limit: int, n_examples: int, rng: random.Random) -> dict:
    counts = collections.Counter()
    examples: dict[str, list[tuple]] = collections.defaultdict(list)
    seen: dict[str, int] = collections.Counter()
    pattern_hits = collections.Counter()
    verified = collections.Counter()
    scanned = 0
    empty_text = 0

    print(f"\nStreaming reviews: raw_review_{category}, target {limit:,}")
    for record in iter_records(f"raw_review_{category}", category, limit):
        scanned += 1
        title = record.get("title") or ""
        text = record.get("text") or ""
        if not text.strip():
            empty_text += 1
        if record.get("verified_purchase"):
            verified["verified"] += 1

        label, hits = label_fit(title, text)
        counts[label] += 1
        if label != "none":
            for bucket, pattern in hits.items():
                pattern_hits[(bucket, pattern)] += 1
        if label == "none":
            continue

        # Reservoir sample so examples are not all drawn from the file head.
        seen[label] += 1
        key = label
        item = (title, text, hits)
        if len(examples[key]) < n_examples:
            examples[key].append(item)
        else:
            j = rng.randrange(seen[key])
            if j < n_examples:
                examples[key][j] = item

        if scanned % 25000 == 0:
            print(f"  ... {scanned:,} reviews scanned")

    return {
        "scanned": scanned,
        "empty_text": empty_text,
        "verified": verified["verified"],
        "counts": counts,
        "examples": examples,
        "pattern_hits": pattern_hits,
    }


# ---------------------------------------------------------------------------
# Question 2
# ---------------------------------------------------------------------------

def normalise_size(raw: str) -> tuple[str | None, str]:
    """Conservatively normalise a free-text size string.

    Returns (normalised, reason). `normalised` is None when the string cannot be
    mapped without guessing; `reason` records which route succeeded or failed.
    """
    if raw is None:
        return None, "absent"
    value = WS_RX.sub(" ", str(raw)).strip().lower()
    if not value:
        return None, "empty"

    # Strip a trailing parenthetical measurement gloss: XL (Chest 44-46")
    stripped = re.sub(r"\s*\([^)]*\)\s*$", "", value).strip()
    candidate = stripped or value
    candidate = candidate.strip(" .,-/")

    if ONE_SIZE.match(candidate):
        return None, "one_size"

    if candidate in ALPHA_SIZE_LADDER:
        return ALPHA_SIZE_LADDER[candidate], "alpha"

    # "us large", "size large", "large / l"
    for prefix in ("size ", "us ", "usa ", "u.s. "):
        if candidate.startswith(prefix):
            tail = candidate[len(prefix):].strip()
            if tail in ALPHA_SIZE_LADDER:
                return ALPHA_SIZE_LADDER[tail], "alpha"

    waist = WAIST_INSEAM.match(candidate)
    if waist:
        return f"W{waist.group(1)}xL{waist.group(2)}", "waist_inseam"

    numeric = PLAIN_NUMERIC.match(candidate)
    if numeric:
        return f"NUM{numeric.group(1)}", "numeric"

    return None, "unmapped"


def probe_meta(category: str, limit: int, n_examples: int, rng: random.Random) -> dict:
    scanned = 0
    details_present = 0
    key_freq = collections.Counter()
    strict_hit = 0
    loose_hit = 0
    loose_keys = collections.Counter()
    norm_reasons = collections.Counter()
    norm_values = collections.Counter()
    raw_samples: list[str] = []
    seen_raw = 0

    gender_counts = collections.Counter()
    half_counts = collections.Counter()
    cells = collections.Counter()
    gradient_cells = collections.Counter()
    categories_present = 0
    both_recovered = 0
    size_and_cell = 0

    print(f"\nStreaming item metadata: raw_meta_{category}, target {limit:,}")
    for record in iter_records(f"raw_meta_{category}", category, limit):
        scanned += 1
        size_key = None
        normalised_size = None

        # --- Question 2: size in details -------------------------------------
        details = record.get("details")
        if isinstance(details, str):
            try:
                details = json.loads(details)
            except (ValueError, TypeError):
                details = None
        if isinstance(details, dict) and details:
            details_present += 1
            normalised_keys = {}
            for key in details:
                flat = WS_RX.sub(" ", str(key)).strip().lower().rstrip(":")
                key_freq[flat] += 1
                normalised_keys[flat] = key

            size_key = None
            for flat in normalised_keys:
                if SIZE_KEY_STRICT.match(flat):
                    size_key = normalised_keys[flat]
                    strict_hit += 1
                    break
            if size_key is None:
                for flat, original in normalised_keys.items():
                    if SIZE_KEY_LOOSE.search(flat) and not SIZE_KEY_LOOSE_EXCLUDE.search(flat):
                        size_key = original
                        loose_keys[flat] += 1
                        loose_hit += 1
                        break

            if size_key is not None:
                raw = details[size_key]
                normalised_size, reason = normalise_size(raw)
                norm_reasons[reason] += 1
                if normalised_size:
                    norm_values[normalised_size] += 1
                seen_raw += 1
                display = f"{size_key} = {raw!r}"
                if len(raw_samples) < n_examples * 6:
                    raw_samples.append(display)
                else:
                    j = rng.randrange(seen_raw)
                    if j < len(raw_samples):
                        raw_samples[j] = display
            else:
                norm_reasons["no_size_key"] += 1
        else:
            norm_reasons["no_details"] += 1

        # --- Question 3: gender and body half from categories ----------------
        cats = record.get("categories") or []
        if isinstance(cats, str):
            cats = [cats]
        segments = [WS_RX.sub(" ", str(c)).strip().lower() for c in cats if c]
        if segments:
            categories_present += 1
        garment_segments = [s for s in segments if s not in ROOT_SEGMENTS]
        joined_all = " | ".join(segments)

        gender = classify_gender(joined_all)
        half = classify_half(garment_segments)
        gender_counts[gender] += 1
        half_counts[half] += 1

        if gender in ("men", "women") and half in ("upper", "lower"):
            both_recovered += 1
            cells[(gender, half)] += 1
            grad = classify_gradient(garment_segments)
            if grad:
                gradient_cells[(gender, grad)] += 1
            if normalised_size is not None:
                size_and_cell += 1

        if scanned % 10000 == 0:
            print(f"  ... {scanned:,} items scanned")

    return {
        "scanned": scanned,
        "details_present": details_present,
        "key_freq": key_freq,
        "strict_hit": strict_hit,
        "loose_hit": loose_hit,
        "loose_keys": loose_keys,
        "norm_reasons": norm_reasons,
        "norm_values": norm_values,
        "raw_samples": raw_samples,
        "gender_counts": gender_counts,
        "half_counts": half_counts,
        "cells": cells,
        "gradient_cells": gradient_cells,
        "categories_present": categories_present,
        "both_recovered": both_recovered,
        "size_and_cell": size_and_cell,
    }


def classify_gender(joined: str) -> str:
    if CHILD_RX.search(joined):
        return "children_excluded"
    men = bool(GENDER_RX["men"].search(joined))
    women = bool(GENDER_RX["women"].search(joined))
    if men and women:
        return "ambiguous"
    if men:
        return "men"
    if women:
        return "women"
    if UNISEX_RX.search(joined):
        return "unisex"
    return "unknown"


def _classify_leaf_first(segments: list[str], rules: dict[str, re.Pattern]) -> str | None:
    """Match `rules` against category segments, most specific segment first.

    Amazon's taxonomy puts multi-class parent nodes above single-class leaves --
    "women | tops, tees & blouses | blouses & button-down shirts". Matching the
    joined path lets the parent's "tees" beat the leaf's "blouses", which
    misfiled 479 women's button-downs as t-shirts in the 2026-08-08 run. So walk
    from the leaf outwards, and treat any segment naming two or more classes as
    ambiguous at that level and keep walking.
    """
    for segment in reversed(segments):
        matched = [label for label, rx in rules.items() if rx.search(segment)]
        if len(matched) == 1:
            return matched[0]
    return None


HALF_RULES = {"excluded": EXCLUDED_RX, "lower": LOWER_RX, "upper": UPPER_RX}


def classify_half(segments: list[str]) -> str:
    if not segments:
        return "unknown"
    return _classify_leaf_first(segments, HALF_RULES) or "unknown"


def _gradient_of_segment(segment: str) -> str | None:
    """Return the gradient class a single category segment names, or None.

    None covers both "names nothing" and "names two classes at once" -- a node
    like "tops, tees & blouses" is genuinely mixed and must not be forced into
    one arm of the primary comparison.
    """
    classes = set()
    if GRADIENT_JEANS_RX.search(segment):
        classes.add("jeans_trousers")
    if GRADIENT_TEE_RX.search(segment):
        classes.add("tshirt")
    if GRADIENT_SHIRT_RX.search(GRADIENT_TEE_STRIP_RX.sub(" ", segment)):
        classes.add("shirt")
    return classes.pop() if len(classes) == 1 else None


def classify_gradient(segments: list[str]) -> str | None:
    if not segments:
        return None
    if _classify_leaf_first(segments, HALF_RULES) == "excluded":
        return None
    for segment in reversed(segments):
        found = _gradient_of_segment(segment)
        if found:
            return found
    return None


# ---------------------------------------------------------------------------
# Precision sample (DESIGN.md 4.1 -- the binding gate measurement)
# ---------------------------------------------------------------------------

PRECISION_COLUMNS = [
    "review_id_hash", "asin", "parent_asin", "product_title", "category_path",
    "gender", "body_half", "review_title", "review_text", "assigned_bucket",
    "human_label",
]


def flatten_categories(cats) -> tuple[list[str], list[str]]:
    """Return (all_segments, garment_segments), lowercased and whitespace-normalised.

    Mirrors the flattening in probe_meta exactly: gender is read from the whole
    path, body half only from segments below the taxonomy root, because the root
    ("clothing, shoes & jewelry") names no body half.
    """
    if isinstance(cats, str):
        cats = [cats]
    segments = [WS_RX.sub(" ", str(c)).strip().lower() for c in (cats or []) if c]
    return segments, [s for s in segments if s not in ROOT_SEGMENTS]


def _review_year(record: dict) -> int | None:
    raw = record.get("timestamp")
    if raw is None:
        return None
    try:
        value = float(raw)
    except (TypeError, ValueError):
        return None
    if value > 1e11:
        value /= 1000.0
    try:
        return datetime.datetime.utcfromtimestamp(value).year
    except (OverflowError, OSError, ValueError):
        return None


def review_id_hash(record: dict) -> str:
    """Stable pseudonymous id. The raw user_id never leaves this function.

    DESIGN.md 6 requires hashing user identifiers at ingest. The triple below is
    unique per review in practice and is not reversible to a user_id.
    """
    raw = "|".join(str(record.get(k, "")) for k in ("user_id", "asin", "timestamp"))
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()[:16]


def build_style_index(category: str, limit: int) -> tuple[dict, dict]:
    """parent_asin -> (gender, body_half, category_path) for in-scope styles only.

    In scope means the item resolves to men or women AND to upper or lower under
    the DESIGN.md 1.3 assignment. Everything else is dropped here rather than
    later, so the sample is garment-scoped by construction.
    """
    index: dict[str, tuple[str, str, str, str]] = {}
    stats = {"seen": 0, "in_scope": 0}
    for record in iter_records(f"raw_meta_{category}", category, limit):
        stats["seen"] += 1
        segments, garment_segments = flatten_categories(record.get("categories"))
        if not segments:
            continue
        gender = classify_gender(" | ".join(segments))
        if gender not in ("men", "women"):
            continue
        half = classify_half(garment_segments)
        if half not in ("upper", "lower"):
            continue
        parent = record.get("parent_asin")
        if not parent:
            continue
        stats["in_scope"] += 1
        title = WS_RX.sub(" ", str(record.get("title") or "")).strip()
        index[parent] = (gender, half, " | ".join(segments), title)
    return index, stats


def draw_precision_sample(category: str, review_limit: int, index: dict,
                          per_stratum: int, rng: random.Random,
                          window_from: int = 0) -> tuple[list, dict]:
    """Reservoir-sample `per_stratum` reviews for each (bucket, gender) cell.

    Reservoir sampling gives every eligible review an equal probability of
    selection in a single pass, without holding the stream in memory and without
    needing to know the eligible count in advance.
    """
    strata = [(b, g) for b in FIT_DICTIONARY for g in ("men", "women")]
    reservoir: dict[tuple[str, str], list] = {s: [] for s in strata}
    seen_count: dict[tuple[str, str], int] = {s: 0 for s in strata}
    stats = {"scanned": 0, "joined": 0, "labelled": 0, "ambiguous": 0,
             "outside_window": 0}

    for record in iter_records(f"raw_review_{category}", category, review_limit):
        stats["scanned"] += 1
        if window_from:
            # DESIGN.md 5.8: measure the dictionary on text the analysis will
            # actually see. Reviews outside the analysis window are not sampled.
            year = _review_year(record)
            if year is None or year < window_from:
                stats["outside_window"] += 1
                continue
        parent = record.get("parent_asin")
        entry = index.get(parent) if parent else None
        if entry is None:
            continue
        stats["joined"] += 1
        bucket, _hits = label_fit(record.get("title", ""), record.get("text", ""))
        if bucket == "ambiguous":
            stats["ambiguous"] += 1
            continue
        if bucket not in FIT_DICTIONARY:
            continue
        gender, half, path, product_title = entry
        stats["labelled"] += 1
        key = (bucket, gender)
        seen_count[key] += 1
        row = {
            "review_id_hash": review_id_hash(record),
            "asin": record.get("asin", ""),
            "parent_asin": parent,
            "product_title": product_title,
            "category_path": path,
            "gender": gender,
            "body_half": half,
            "review_title": WS_RX.sub(" ", str(record.get("title", ""))).strip(),
            "review_text": WS_RX.sub(" ", str(record.get("text", ""))).strip(),
            "assigned_bucket": bucket,
            "human_label": "",
        }
        pool = reservoir[key]
        if len(pool) < per_stratum:
            pool.append(row)
        else:
            j = rng.randrange(seen_count[key])
            if j < per_stratum:
                pool[j] = row

    rows = [r for s in strata for r in reservoir[s]]
    stats["eligible_per_stratum"] = {f"{b}/{g}": seen_count[(b, g)] for b, g in strata}
    stats["drawn_per_stratum"] = {f"{b}/{g}": len(reservoir[(b, g)]) for b, g in strata}
    return rows, stats


# What the labeller sees. `product_title` is deliberately NOT blinded: the coding
# rules require non-garment items to be marked `none`, and that rule cannot be
# applied to review text alone -- "too big, returned it" is unclassifiable
# without knowing whether the product is a t-shirt or a watch strap. The 0/4
# ran_large precision on Amazon_Fashion was caused entirely by a purse, two watch
# straps and a pair of glasses, so this is the error mode the sample exists to
# measure. A product title carries no signal about WHICH BUCKET the dictionary
# assigned, so it does not compromise the blind. It may leak gender, which is the
# weaker cost: gender is a secondary breakdown, not the gate.
# Column set fixed by docs/coding-guide.md v1.0 §9. Two changes from the first
# emission, both from the coding guide: `buyer_gender_mismatch` is renamed
# `wearer_gender_mismatch` because what matters is whose body the judgement
# describes, not who paid (guide §2); and `calibration_stated` is added, which
# flags reviews attributing the sizing to the brand or a regional convention
# (guide §8) and gives a cheap manual upper bound on the DESIGN.md 5.9 confound.
BLIND_COLUMNS = ["review_id_hash", "product_title", "review_title", "review_text",
                 "human_label", "wearer_gender_mismatch", "calibration_stated"]


def emit_precision_sample(path: pathlib.Path, rows: list, rng: random.Random) -> None:
    """Write the key file and the BLIND labelling file.

    An unblinded precision measurement is not a measurement: a labeller who can
    see `assigned_bucket` is scoring their agreement with a number already in
    front of them. The blind file therefore carries only the review id and the
    text. Gender, body half, category path and the assigned bucket all stay in
    the key file, which the labeller does not open, and are re-joined on
    `review_id_hash` after the labels come back.
    """
    rng.shuffle(rows)  # so row order carries no information about the bucket
    path.parent.mkdir(parents=True, exist_ok=True)

    key_path = path.with_name(path.stem + "_key.csv")
    with key_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=PRECISION_COLUMNS)
        writer.writeheader()
        writer.writerows(rows)
    print(f"  [key  ] wrote {key_path}  (NOT for labelling -- contains assigned_bucket)")

    blind = [{"review_id_hash": row["review_id_hash"],
              "product_title": row["product_title"],
              "review_title": row["review_title"],
              "review_text": row["review_text"],
              "human_label": "",
              "wearer_gender_mismatch": "",
              "calibration_stated": ""} for row in rows]

    blind_path = path.with_name(path.stem + "_blind.csv")
    with blind_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=BLIND_COLUMNS)
        writer.writeheader()
        writer.writerows(blind)
    _emit_xlsx(blind_path.with_suffix(".xlsx"), blind, BLIND_COLUMNS)
    print(f"  [blind] wrote {blind_path} and .xlsx  <- THIS is the labelling file")


def _emit_xlsx(path: pathlib.Path, rows: list, columns: list | None = None) -> None:
    """Same rows, same columns, as a workbook.

    Review text contains commas, quotes and newlines, which is exactly what a
    hand-labeller in a spreadsheet should not have to fight. The xlsx is the
    labelling surface; the csv stays as the machine-readable copy. Both are
    git-ignored permanently -- they carry raw review text (DESIGN.md 6).
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.worksheet.datavalidation import DataValidation
    except ImportError:
        print("  [xlsx] openpyxl not installed; wrote csv only  (pip install openpyxl)")
        return

    columns = columns or PRECISION_COLUMNS
    book = Workbook()
    sheet = book.active
    sheet.title = "precision_sample"
    sheet.append(columns)

    header_fill = PatternFill("solid", fgColor="DDDDDD")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    sheet.freeze_panes = "A2"

    for row in rows:
        sheet.append([row[column] for column in columns])

    widths = {"review_id_hash": 18, "asin": 12, "parent_asin": 13,
              "product_title": 52, "category_path": 46, "gender": 8, "body_half": 10,
              "review_title": 40, "review_text": 90,
              "assigned_bucket": 15, "human_label": 16,
              "wearer_gender_mismatch": 24, "calibration_stated": 20}
    for index, column in enumerate(columns, start=1):
        letter = sheet.cell(row=1, column=index).column_letter
        sheet.column_dimensions[letter].width = widths.get(column, 14)
        if column in ("review_title", "review_text", "category_path", "product_title"):
            for cell in sheet[letter][1:]:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Constrain the label column so the returned file is machine-readable.
    label_index = columns.index("human_label") + 1
    label_letter = sheet.cell(row=1, column=label_index).column_letter
    validation = DataValidation(
        type="list",
        formula1='"ran_small,true_to_size,ran_large,none,unclear"',
        allow_blank=False,
        showDropDown=False,
    )
    validation.prompt = ("Relative to the buyer's own body, did the garment they actually "
                         "RECEIVED run small, fit, or run large? Code physical fit, not "
                         "satisfaction. 'none' = no fit judgement at all; 'unclear' = fit "
                         "discussed but direction undeterminable. See docs/coding-guide.md.")
    validation.promptTitle = "human_label"
    sheet.add_data_validation(validation)
    validation.add(f"{label_letter}2:{label_letter}{len(rows) + 1}")

    # Flag columns, per docs/coding-guide.md §9. Leave nothing blank: "no" and
    # "unclear" are answers, so allow_blank is False on both.
    for name, values, prompt in [
        ("wearer_gender_mismatch", '"yes,no,unclear"',
         "yes when the fit judgement is given for a body whose gender differs from the "
         "product's gender. Buying on someone else's behalf is not by itself a mismatch."),
        ("calibration_stated", '"yes,no"',
         "yes when the review explicitly attributes the sizing to the brand, the "
         "manufacturer, or a regional sizing convention."),
    ]:
        if name not in columns:
            continue
        flag_index = columns.index(name) + 1
        flag_letter = sheet.cell(row=1, column=flag_index).column_letter
        flag_validation = DataValidation(type="list", formula1=values,
                                         allow_blank=False, showDropDown=False)
        flag_validation.prompt = prompt
        flag_validation.promptTitle = name
        sheet.add_data_validation(flag_validation)
        flag_validation.add(f"{flag_letter}2:{flag_letter}{len(rows) + 1}")

    book.save(path)
    print(f"  [xlsx] wrote {path}")


# ---------------------------------------------------------------------------
# Reporting
# ---------------------------------------------------------------------------

def pct(n: int, d: int) -> str:
    return f"{100.0 * n / d:6.2f}%" if d else "   n/a"


def rule(title: str) -> None:
    print("\n" + "=" * 78)
    print(title)
    print("=" * 78)


def report_reviews(res: dict, n_examples: int) -> None:
    rule("QUESTION 1 -- fit judgements in review title+text")
    scanned = res["scanned"]
    counts = res["counts"]
    labelled = sum(counts[b] for b in FIT_DICTIONARY)
    print(f"reviews scanned                  {scanned:>9,}")
    print(f"empty review text                {res['empty_text']:>9,}  {pct(res['empty_text'], scanned)}")
    print(f"verified_purchase = True         {res['verified']:>9,}  {pct(res['verified'], scanned)}")
    print()
    print(f"usable fit label (exactly 1)     {labelled:>9,}  {pct(labelled, scanned)}")
    print(f"ambiguous, dropped (>=2 buckets) {counts['ambiguous']:>9,}  {pct(counts['ambiguous'], scanned)}")
    print(f"no fit language                  {counts['none']:>9,}  {pct(counts['none'], scanned)}")
    print()
    print("split within the labelled set:")
    for bucket in FIT_DICTIONARY:
        print(f"  {bucket:<14} {counts[bucket]:>9,}  {pct(counts[bucket], labelled)} of labelled"
              f"   {pct(counts[bucket], scanned)} of scanned")

    print("\ntop firing patterns per bucket (audit trail):")
    for bucket in FIT_DICTIONARY:
        rows = [(p, c) for (b, p), c in res["pattern_hits"].items() if b == bucket]
        rows.sort(key=lambda r: -r[1])
        print(f"  [{bucket}]")
        for pattern, count in rows[:6]:
            print(f"    {count:>7,}  {pattern}")

    print("\n" + "-" * 78)
    print("HAND-VERIFICATION SAMPLE -- read these and score each as correct or not.")
    print("-" * 78)
    for bucket in list(FIT_DICTIONARY) + ["ambiguous"]:
        print(f"\n### {bucket}  (n shown = {len(res['examples'].get(bucket, []))})")
        for i, (title, text, hits) in enumerate(res["examples"].get(bucket, []), 1):
            pattern = next(iter(hits.values()))
            print(f"  {i}. matched: {' + '.join(hits) if len(hits) > 1 else pattern}")
            print(f"     {snippet(title, text, pattern)}")


def report_meta(res: dict) -> None:
    scanned = res["scanned"]

    rule("QUESTION 2 -- purchased size in the item `details` field")
    print(f"items scanned                    {scanned:>9,}")
    print(f"non-empty `details` dict         {res['details_present']:>9,}  {pct(res['details_present'], scanned)}")
    size_any = res["strict_hit"] + res["loose_hit"]
    print(f"details has exact 'size' key     {res['strict_hit']:>9,}  {pct(res['strict_hit'], scanned)}")
    print(f"details has other size-ish key   {res['loose_hit']:>9,}  {pct(res['loose_hit'], scanned)}")
    print(f"details exposes any size         {size_any:>9,}  {pct(size_any, scanned)}")

    norm = res["norm_reasons"]
    normalisable = norm["alpha"] + norm["waist_inseam"] + norm["numeric"]
    print()
    print("conservative normalisation of the recovered size strings:")
    for reason in ("alpha", "waist_inseam", "numeric", "one_size", "unmapped", "empty"):
        if norm[reason]:
            print(f"  {reason:<14} {norm[reason]:>9,}  {pct(norm[reason], size_any)} of recovered")
    print(f"  {'NORMALISABLE':<14} {normalisable:>9,}  {pct(normalisable, size_any)} of recovered"
          f"   {pct(normalisable, scanned)} of all items")
    dropped = size_any - normalisable
    print(f"  {'dropped':<14} {dropped:>9,}  {pct(dropped, size_any)} of recovered")

    print("\ntop 20 `details` keys observed (is a size key even present?):")
    for key, count in res["key_freq"].most_common(20):
        print(f"  {count:>8,}  {key}")
    if res["loose_keys"]:
        print("\nsize-ish keys matched by the loose rule:")
        for key, count in res["loose_keys"].most_common(10):
            print(f"  {count:>8,}  {key}")

    print("\nraw size strings, as written by sellers (DESIGN.md 5.3 check):")
    for sample in res["raw_samples"][:24]:
        print(f"  {sample}")
    if res["norm_values"]:
        print("\nmost common normalised values:")
        for value, count in res["norm_values"].most_common(15):
            print(f"  {count:>8,}  {value}")

    rule("QUESTION 3 -- gender and body half from `categories` (DESIGN.md 1.3)")
    print(f"items scanned                    {scanned:>9,}")
    print(f"non-empty `categories`           {res['categories_present']:>9,}  {pct(res['categories_present'], scanned)}")
    print("\ngender recovery:")
    for key, count in res["gender_counts"].most_common():
        print(f"  {key:<20} {count:>9,}  {pct(count, scanned)}")
    print("\nbody-half recovery:")
    for key, count in res["half_counts"].most_common():
        print(f"  {key:<20} {count:>9,}  {pct(count, scanned)}")
    print()
    print(f"BOTH gender and half recovered   {res['both_recovered']:>9,}  {pct(res['both_recovered'], scanned)}")
    print(f"  ... and size normalisable      {res['size_and_cell']:>9,}  {pct(res['size_and_cell'], scanned)}")

    cells = res["cells"]
    print("\nthe four gender x half cells (item counts, wide DESIGN.md 1.3 sets):")
    print(f"  {'':<8}{'upper':>12}{'lower':>12}")
    for gender in ("men", "women"):
        print(f"  {gender:<8}{cells[(gender, 'upper')]:>12,}{cells[(gender, 'lower')]:>12,}")
    men_lower = cells[("men", "lower")]
    total_cells = sum(cells.values())
    print()
    print(f"MEN'S-LOWER CELL (DESIGN.md 5.5, the binding constraint): {men_lower:,} items"
          f"  = {pct(men_lower, total_cells)} of the four cells")

    grad = res["gradient_cells"]
    print("\nprimary three-step gradient of DESIGN.md 1.2 (item counts):")
    print(f"  {'':<8}{'tshirt':>12}{'shirt':>12}{'jeans/trs':>12}")
    for gender in ("men", "women"):
        print(f"  {gender:<8}{grad[(gender, 'tshirt')]:>12,}"
              f"{grad[(gender, 'shirt')]:>12,}{grad[(gender, 'jeans_trousers')]:>12,}")


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__.split("\n")[0])
    parser.add_argument("--reviews", type=int, default=50_000,
                        help="number of reviews to stream (default 50000)")
    parser.add_argument("--items", type=int, default=30_000,
                        help="number of item metadata records to stream (default 30000)")
    parser.add_argument("--category", default="Amazon_Fashion",
                        help="Amazon Reviews'23 category (default Amazon_Fashion)")
    parser.add_argument("--examples", type=int, default=4,
                        help="example matches printed per bucket (default 4)")
    parser.add_argument("--seed", type=int, default=20260808,
                        help="seed for example sampling, for reproducibility")
    parser.add_argument("--spread", type=int, default=0,
                        help="block-sample across N disjoint file offsets "
                             "instead of reading a prefix (0 = prefix, biased)")
    parser.add_argument("--skip-reviews", action="store_true")
    parser.add_argument("--skip-meta", action="store_true")
    parser.add_argument("--precision-sample", action="store_true",
                        help="draw the DESIGN.md 4.1 hand-labelling sample and exit")
    parser.add_argument("--sample-items", type=int, default=200_000,
                        help="meta records streamed to build the style index (default 200000)")
    parser.add_argument("--sample-reviews", type=int, default=600_000,
                        help="reviews streamed for the precision sample (default 600000)")
    parser.add_argument("--per-stratum", type=int, default=50,
                        help="reviews drawn per bucket x gender cell (default 50 -> 300 rows)")
    parser.add_argument("--sample-out", default="data/processed/precision_sample.csv")
    parser.add_argument("--window-from", type=int, default=0,
                        help="only sample reviews from this calendar year onward "
                             "(DESIGN.md 5.8 analysis window)")
    args = parser.parse_args()

    rng = random.Random(args.seed)

    global SPREAD_BLOCKS
    SPREAD_BLOCKS = args.spread

    if args.precision_sample:
        return run_precision_sample(args, rng)

    rule(f"THE VANITY GAP -- Phase 1 Amazon probe -- category {args.category}")
    print(f"reviews requested {args.reviews:,}   items requested {args.items:,}   seed {args.seed}")
    print("stream-only: no file is downloaded and nothing is written to disk")

    if not args.skip_reviews:
        report_reviews(probe_reviews(args.category, args.reviews, args.examples, rng),
                       args.examples)
    if not args.skip_meta:
        report_meta(probe_meta(args.category, args.items, args.examples, rng))

    rule("END OF PROBE -- apply the DESIGN.md 4.1 gate to these numbers")
    return 0


def run_precision_sample(args, rng: random.Random) -> int:
    rule(f"PRECISION SAMPLE -- category {args.category}")
    print(f"seed {args.seed}   per-stratum target {args.per_stratum}")
    print(f"style index from {args.sample_items:,} meta records")
    print(f"reviews streamed  {args.sample_reviews:,}")

    index, meta_stats = build_style_index(args.category, args.sample_items)
    print(f"\nmeta records seen      {meta_stats['seen']:>9,}")
    print(f"in-scope styles indexed {meta_stats['in_scope']:>8,}  "
          f"{pct(meta_stats['in_scope'], meta_stats['seen'])}")

    rows, stats = draw_precision_sample(
        args.category, args.sample_reviews, index, args.per_stratum, rng,
        window_from=args.window_from)

    print(f"\nreviews scanned        {stats['scanned']:>9,}")
    print(f"joined to an in-scope style {stats['joined']:>4,}  "
          f"{pct(stats['joined'], stats['scanned'])}")
    print(f"carrying a single bucket    {stats['labelled']:>4,}  "
          f"{pct(stats['labelled'], stats['joined'])} of joined")
    print(f"ambiguous, dropped          {stats['ambiguous']:>4,}")

    print("\neligible population per stratum (the sampling frame):")
    for key, value in stats["eligible_per_stratum"].items():
        print(f"  {key:<28} {value:>7,}   drawn {stats['drawn_per_stratum'][key]:>4,}")

    if not rows:
        print("\nNO ROWS DRAWN -- the join produced nothing. Not writing a file.")
        return 1

    out = pathlib.Path(args.sample_out)
    emit_precision_sample(out, rows, rng)
    print(f"\nwrote {len(rows):,} rows to {out}")
    print("`human_label` is blank by design -- the repository owner labels it.")
    print("NOT FOR PUBLICATION: this file contains raw review text (DESIGN.md 6).")
    return 0


if __name__ == "__main__":
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.exit(main())

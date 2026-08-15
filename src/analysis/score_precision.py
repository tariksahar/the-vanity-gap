"""Score the dictionary against the blind hand labels.

Joins the blind labelling file to the key file on `review_id_hash` and computes
precision per bucket -- the DESIGN.md 4.1 gate measure.

Two things this does that a naive score would not.

**Wilson intervals, not normal approximation.** At roughly 25 labelled rows per
stratum the normal approximation to a binomial proportion is unreliable near the
boundaries, and the gate threshold (~80%) is near enough to 1.0 for that to
matter. Wilson is well behaved at small n and does not produce intervals that
run past 1.

**`unclear` reported both ways.** A row the labeller could not resolve is not
evidence that the dictionary was wrong, but excluding it silently would flatter
the dictionary. Precision is therefore reported twice: excluding `unclear` from
the denominator (the headline, since an unscoreable row scores nothing), and
counting it as an error (strict). The gap between them is the exposure.

Usage:
    python src/analysis/score_precision.py
"""

from __future__ import annotations

import collections
import csv
import math
import pathlib
import sys

ROOT = pathlib.Path(__file__).resolve().parents[2]
BUCKETS = ["ran_small", "true_to_size", "ran_large"]
Z = 1.959963985


def wilson(successes: int, n: int, z: float = Z) -> tuple[float, float]:
    """Wilson score interval for a binomial proportion."""
    if n == 0:
        return (0.0, 1.0)
    p = successes / n
    denominator = 1 + z * z / n
    centre = (p + z * z / (2 * n)) / denominator
    half = z * math.sqrt(p * (1 - p) / n + z * z / (4 * n * n)) / denominator
    return (max(0.0, centre - half), min(1.0, centre + half))


def load() -> list[dict]:
    """Rows that carry a human label, with the key fields attached."""
    key_path = ROOT / "data/processed/precision_sample_key.csv"
    key = {r["review_id_hash"]: r for r in csv.DictReader(
        key_path.open(encoding="utf-8"))}

    # Canonical labelled file. The blind workbook it was drawn from is superseded:
    # calibration_stated was re-coded 27 -> 3 and one human_label corrected.
    blind_path = ROOT / "data/processed/precision_sample_labelled_FINAL.xlsx"
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("openpyxl required to read the labelled workbook")
        raise

    sheet = load_workbook(blind_path, data_only=True).active
    header = [c.value for c in sheet[1]]
    rows = []
    for values in sheet.iter_rows(min_row=2, values_only=True):
        record = dict(zip(header, values))
        label = (record.get("human_label") or "").strip()
        if not label:
            continue
        entry = key.get(record["review_id_hash"])
        if entry is None:
            continue
        rows.append({
            "assigned": entry["assigned_bucket"],
            "gender": entry["gender"],
            "body_half": entry["body_half"],
            "human": label,
            "mismatch": (record.get("wearer_gender_mismatch") or "").strip().lower(),
            "calibration": (record.get("calibration_stated") or "").strip().lower(),
        })
    return rows


def score(rows: list[dict], label: str = "") -> dict:
    out = {}
    for bucket in BUCKETS:
        subset = [r for r in rows if r["assigned"] == bucket]
        if not subset:
            continue
        correct = sum(1 for r in subset if r["human"] == bucket)
        unclear = sum(1 for r in subset if r["human"] == "unclear")
        none = sum(1 for r in subset if r["human"] == "none")
        other = sum(1 for r in subset
                    if r["human"] in BUCKETS and r["human"] != bucket)
        scoreable = len(subset) - unclear
        out[bucket] = {
            "n": len(subset), "correct": correct, "unclear": unclear,
            "none": none, "other_bucket": other, "scoreable": scoreable,
            "precision": correct / scoreable if scoreable else 0.0,
            "ci": wilson(correct, scoreable),
            "strict": correct / len(subset),
            "strict_ci": wilson(correct, len(subset)),
        }
    return out


def _table(scored: dict, title: str) -> None:
    print(f"\n{title}")
    print(f"{'bucket':<14}{'n':>5}{'ok':>5}{'none':>6}{'othr':>6}{'uncl':>6}"
          f"{'precision':>11}{'95% CI (Wilson)':>22}{'strict':>9}")
    for bucket, s in scored.items():
        low, high = s["ci"]
        print(f"{bucket:<14}{s['n']:>5}{s['correct']:>5}{s['none']:>6}"
              f"{s['other_bucket']:>6}{s['unclear']:>6}"
              f"{s['precision']:>10.1%}   [{low:>5.1%}, {high:>5.1%}]"
              f"{s['strict']:>9.1%}")


def main() -> int:
    rows = load()
    print("=" * 84)
    print("DICTIONARY PRECISION -- blind hand labels")
    print("=" * 84)
    print(f"labelled rows: {len(rows)} of 300 drawn "
          f"({len(rows) / 300:.0%}); labelling stopped by decision")

    _table(score(rows), "ALL LABELLED ROWS")

    for gender in ("men", "women"):
        subset = [r for r in rows if r["gender"] == gender]
        scored = score(subset)
        smallest = min((s["scoreable"] for s in scored.values()), default=0)
        if smallest < 8:
            print(f"\n[{gender}] smallest scoreable bucket is {smallest} rows "
                  "-- too thin to report a rate; suppressed")
            continue
        _table(scored, f"BY GENDER: {gender.upper()}  (n={len(subset)})")

    print("\n" + "=" * 84)
    print("LABELLER FLAGS")
    print("=" * 84)
    for field, positive in (("calibration", "yes"), ("mismatch", "yes")):
        hits = sum(1 for r in rows if r[field] == positive)
        low, high = wilson(hits, len(rows))
        name = {"calibration": "calibration_stated",
                "mismatch": "wearer_gender_mismatch"}[field]
        print(f"  {name:<24}{hits:>4} / {len(rows)}  = {hits / len(rows):>5.1%}"
              f"   [{low:.1%}, {high:.1%}]   LOWER BOUND")
    print("\n  Both are lower bounds: a reviewer affected by seller calibration")
    print("  who does not say so is not counted, and most reviews say nothing")
    print("  about who wore the garment.")

    print("\n" + "=" * 84)
    print("HUMAN LABEL DISTRIBUTION (what the labeller actually saw)")
    print("=" * 84)
    for label, n in collections.Counter(r["human"] for r in rows).most_common():
        print(f"  {label:<16}{n:>5}  {n / len(rows):>6.1%}")

    print("\n" + "=" * 84)
    print("SUPERSEDED: the >= ~80% per-bucket gate was removed 2026-08-14")
    print("=" * 84)
    scored = score(rows)
    for bucket, s in scored.items():
        low, high = s["ci"]
        if low >= 0.80:
            verdict = "PASS -- interval entirely above threshold"
        elif high < 0.80:
            verdict = "FAIL -- interval entirely below threshold"
        else:
            verdict = "INCONCLUSIVE -- interval spans the threshold"
        print(f"  {bucket:<14}{s['precision']:>6.1%}  {verdict}")
    print("\n  At ~25 scoreable rows per bucket the intervals are wide by")
    print("  construction. An inconclusive verdict is a statement about the")
    print("  sample size, not about the dictionary.")
    return 0


if __name__ == "__main__":
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.exit(main())

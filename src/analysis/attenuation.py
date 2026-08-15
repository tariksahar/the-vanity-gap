"""Attenuation of `tau` caused by misclassification in `fit_score`.

Companion to `power.py`. `power.py` says what effect size the design can detect
in the MEASURED outcome; this module says what fraction of the TRUE effect
survives into that measured outcome. Together they give the operative MDE:

    MDE_operative = MDE_design / lambda

---------------------------------------------------------------------------
WHICH CONDITIONAL -- the point on which this module differs from its brief
---------------------------------------------------------------------------

Write `y  = score(assigned)` for what the pipeline computes and
      `y* = score(true)`     for what we want, with
      score(ran_small) = -1, score(true_to_size) = 0, score(ran_large) = +1.

The estimand is a difference in means of `y*`. We estimate a difference in
means of `y`. Under the standard non-differential misclassification assumption

    P(assigned = k | true = j, X) = P(assigned = k | true = j) = M_kj

it follows that

    E[y | X] = sum_j a_j P(true = j | X),   a_j = sum_k score(k) M_kj

so `E[y|X]` is a reweighting of the true-class probabilities by `a_j`, and the
measured difference is `lambda` times the true difference with

    lambda_forward = (a_large - a_small) / 2                          [1]

**`a_j` requires `P(assigned | true)` -- the FORWARD conditional.**

The brief specified instead

    s_k = sum_j P(true = j | assigned = k) score(j),
    lambda_reverse = (s_large - s_small) / 2                          [2]

which uses the REVERSE conditional. These are different quantities and they do
not generally coincide. [2] is the one our sample estimates directly, because
the sample is stratified on the ASSIGNED bucket, so each row of the confusion
matrix is a draw from `P(true | assigned = k)`. [1] is the one that makes
`MDE_operative = MDE_design / lambda` correct, because attenuation of a
misclassified OUTCOME is governed by sensitivity and specificity -- how the
truth maps into the measurement, not the reverse.

Using [2] in `MDE_design / lambda` would be a category error: under the
assumption that makes [2] meaningful (`P(true | assigned, X) = P(true |
assigned)`) one gets `E[y*|X] = c + lambda_reverse * E[y|X]`, i.e. the TRUE
difference equals `lambda_reverse` times the MEASURED one -- the measurement
would be inflating rather than attenuating, and dividing by it would be the
wrong direction.

Both are therefore computed and reported. **`lambda_forward` is operative.**
[1] is recovered from [2] by Bayes, using the assigned-bucket prevalence of the
analysis population:

    M_kj = P(true=j | assigned=k) * w_k / p_j,   p_j = sum_k P(true=j|assigned=k) w_k

---------------------------------------------------------------------------
AFFINITY RESIDUAL
---------------------------------------------------------------------------

With three classes, `a_j = alpha + lambda * score(j)` is three equations in two
unknowns and does not hold exactly. [1] pins `lambda` on the two extremes. The
middle class then deviates by

    residual = a_0 - (a_small + a_large) / 2

which is reported as a diagnostic: a large residual means the single-number
summary is hiding structure, and the correction should not be trusted to more
than its leading digit.

---------------------------------------------------------------------------
UNCERTAINTY
---------------------------------------------------------------------------

Confidence intervals come from a **cluster bootstrap over `parent_asin`**, not
from Wilson intervals. In the labelled sample four products carry more than 5%
of rows each and the largest carries 10.1%, so rows are not independent and an
interval that assumes they are is too narrow. Parents are resampled globally
rather than within stratum, because 13 parents span more than one assigned
bucket; stratum sizes therefore vary across replicates, which is the
conservative choice.
"""

from __future__ import annotations

import collections
import csv
import pathlib
import random
import sys

ROOT = pathlib.Path(__file__).resolve().parents[2]

BUCKETS = ("ran_small", "true_to_size", "ran_large")
SCORE = {"ran_small": -1.0, "true_to_size": 0.0, "ran_large": 1.0}

# Assigned-bucket prevalence of the ANALYSIS population: labelled reviews joined
# to an in-scope style inside the 2019 window, from the sampling frame of the
# 2026-08-11 draw (docs/phase1-amazon-probe.md 5.9). Not the sample's own
# proportions -- the sample is stratified with equal allocation, so its
# proportions carry no information about the corpus.
ASSIGNED_PREVALENCE = {"ran_small": 1054.0, "true_to_size": 1706.0, "ran_large": 471.0}

LABELLED = ROOT / "data/processed/precision_sample_labelled_FINAL.xlsx"
KEY = ROOT / "data/processed/precision_sample_key.csv"


# ---------------------------------------------------------------------------
# Data
# ---------------------------------------------------------------------------

STORE_MAP = ROOT / "data/processed/parent_store.json"


def load_store_map() -> dict[str, str]:
    """parent_asin -> store, resolved from item metadata. Empty if absent."""
    import json

    if not STORE_MAP.exists():
        return {}
    return json.loads(STORE_MAP.read_text(encoding="utf-8"))


def load_rows(path: pathlib.Path = LABELLED, key_path: pathlib.Path = KEY) -> list[dict]:
    """Labelled rows with the key fields attached. One dict per review."""
    from openpyxl import load_workbook

    store_map = load_store_map()
    key = {r["review_id_hash"]: r for r in csv.DictReader(
        key_path.open(encoding="utf-8"))}
    sheet = load_workbook(path, data_only=True).active
    header = [c.value for c in sheet[1]]
    rows = []
    for values in sheet.iter_rows(min_row=2, values_only=True):
        record = dict(zip(header, values))
        label = (record.get("human_label") or "").strip()
        if not label:
            continue
        entry = key.get(record["review_id_hash"])
        if entry is None:
            continue
        rows.append({
            "parent": entry["parent_asin"],
            "store": store_map.get(entry["parent_asin"]),
            "assigned": entry["assigned_bucket"],
            "gender": entry["gender"],
            "half": entry["body_half"],
            "human": label,
        })
    return rows


def confusion(rows: list[dict]) -> dict[tuple[str, str], int]:
    """(assigned, human) -> count. Raw counts, no normalisation."""
    return collections.Counter((r["assigned"], r["human"]) for r in rows)


# ---------------------------------------------------------------------------
# Lambda
# ---------------------------------------------------------------------------

def _row_conditionals(matrix, none_as_zero: bool) -> dict[str, dict[str, float]] | None:
    """P(true = j | assigned = k), one distribution per assigned bucket.

    `unclear` is always dropped -- it carries no information about the truth.
    `none` means the human found no fit judgement at all; its true score is not
    defined. Treated as 0 when `none_as_zero`, dropped otherwise.
    """
    out = {}
    for k in BUCKETS:
        counts = {j: float(matrix.get((k, j), 0)) for j in BUCKETS}
        if none_as_zero:
            counts["true_to_size"] += float(matrix.get((k, "none"), 0))
        total = sum(counts.values())
        if total == 0:
            return None
        out[k] = {j: c / total for j, c in counts.items()}
    return out


def lambda_reverse(matrix, none_as_zero: bool = True) -> float | None:
    """(s_large - s_small) / 2 with s_k = E[score(true) | assigned = k].

    The brief's formula. Reported for comparison; see the module docstring for
    why it is not the operative one.
    """
    p = _row_conditionals(matrix, none_as_zero)
    if p is None:
        return None
    s = {k: sum(p[k][j] * SCORE[j] for j in BUCKETS) for k in BUCKETS}
    return (s["ran_large"] - s["ran_small"]) / 2.0


def lambda_forward(matrix, prevalence=None, none_as_zero: bool = True):
    """Operative attenuation factor, plus the affinity residual.

    Returns (lambda, residual, a) or (None, None, None) if a class is empty.
    """
    p = _row_conditionals(matrix, none_as_zero)
    if p is None:
        return None, None, None
    prevalence = prevalence or ASSIGNED_PREVALENCE
    total_w = sum(prevalence[k] for k in BUCKETS)
    w = {k: prevalence[k] / total_w for k in BUCKETS}

    # p_j = P(true = j) in the analysis population
    p_true = {j: sum(w[k] * p[k][j] for k in BUCKETS) for j in BUCKETS}
    if any(v <= 0 for v in p_true.values()):
        return None, None, None

    # a_j = E[score(assigned) | true = j] via Bayes
    a = {j: sum(SCORE[k] * w[k] * p[k][j] for k in BUCKETS) / p_true[j]
         for j in BUCKETS}
    lam = (a["ran_large"] - a["ran_small"]) / 2.0
    residual = a["true_to_size"] - (a["ran_small"] + a["ran_large"]) / 2.0
    return lam, residual, a


# P(assigned = k | gender, body_half) in the analysis population. Measured by
# `style_definition_probe.py` on 3,000,000 block-sampled reviews in the 2019
# window against a 400,000-record style index -- 16,029 labelled observations.
# No earlier probe reported this breakdown; `p_0(cell)` cannot be built without
# it. Counts, not shares, so the weighting is transparent.
CELL_ASSIGNED: dict[tuple[str, str], dict[str, float]] = {
    ("men", "upper"):   {"ran_small": 826.0, "true_to_size": 1217.0, "ran_large": 342.0},
    ("men", "lower"):   {"ran_small": 546.0, "true_to_size": 1113.0, "ran_large": 346.0},
    ("women", "upper"): {"ran_small": 3358.0, "true_to_size": 4025.0, "ran_large": 1082.0},
    ("women", "lower"): {"ran_small": 947.0, "true_to_size": 1693.0, "ran_large": 534.0},
}


def p_zero_by_cell(matrix, cell_assigned=None, none_as_zero: bool = True):
    """P(true = true_to_size | cell), for each gender x half cell.

    Built as `sum_k P(true=0 | assigned=k) * P(assigned=k | cell)`. The first
    factor comes from the 149 hand labels; the second from the whole analysis
    population. So the uncertainty in `p_0` is essentially all on the label side.
    """
    cell_assigned = cell_assigned or CELL_ASSIGNED
    if not cell_assigned:
        return {}
    p = _row_conditionals(matrix, none_as_zero)
    if p is None:
        return {}
    out = {}
    for cell, assigned in cell_assigned.items():
        total = sum(assigned.values())
        if total <= 0:
            continue
        out[cell] = sum(p[k]["true_to_size"] * assigned[k] / total for k in BUCKETS)
    return out


def sd_fit_score(prevalence=None) -> float:
    """SD of the MEASURED fit_score in the analysis population.

    Scores are -1 / 0 / +1 with the assigned-bucket prevalence, so
    Var = (w_small + w_large) - (w_large - w_small)^2. The bias term is a
    difference of means and is converted to SD units by dividing by this.
    """
    prevalence = prevalence or ASSIGNED_PREVALENCE
    total = sum(prevalence[k] for k in BUCKETS)
    w = {k: prevalence[k] / total for k in BUCKETS}
    mean = w["ran_large"] - w["ran_small"]
    return ((w["ran_small"] + w["ran_large"]) - mean ** 2) ** 0.5


def residual_bias(rows: list[dict], cell_assigned=None, prevalence=None):
    """The term the residual contributes to `tau`, in SD units of fit_score.

    From `E[y | cell] = c + lambda * E[y* | cell] + delta * p_0(cell)`:

        tau_measured = [l_m*D*_m + d_m*Dp0_m] - [l_w*D*_w + d_w*Dp0_w]

    `c` cancels in the double difference. `delta * Dp0` does not. So even with
    `lambda_men == lambda_women`, a difference in `delta` biases `tau` whenever
    `Dp0 != 0`.

    Returns (bias_in_sd, parts) or (None, {}) when `CELL_ASSIGNED` is unset.
    """
    cell_assigned = cell_assigned or CELL_ASSIGNED
    if not cell_assigned:
        return None, {}
    parts = {}
    for gender in ("men", "women"):
        subset = [r for r in rows if r["gender"] == gender]
        _lam, delta, _a = lambda_forward(confusion(subset), prevalence)
        if delta is None:
            return None, {}
        p0 = p_zero_by_cell(confusion(subset), cell_assigned)
        upper, lower = p0.get((gender, "upper")), p0.get((gender, "lower"))
        if upper is None or lower is None:
            return None, {}
        parts[gender] = {"delta": delta, "p0_upper": upper, "p0_lower": lower,
                         "dp0": upper - lower, "term": delta * (upper - lower)}
    raw = parts["men"]["term"] - parts["women"]["term"]
    return raw / sd_fit_score(prevalence), parts


def mde_operative(mde_design: float, lam: float) -> float:
    """The true effect size the study can actually detect."""
    if lam <= 0:
        raise ValueError("lambda must be positive; a non-positive lambda means "
                         "the measurement carries no signal about the truth")
    return mde_design / lam


# ---------------------------------------------------------------------------
# Cluster bootstrap
# ---------------------------------------------------------------------------

def cluster_bootstrap(rows: list[dict], statistic, replicates: int = 4000,
                      seed: int = 20260814,
                      unit: str = "parent") -> tuple[float, float, list[float]]:
    """Percentile CI for `statistic(rows)`, resampling clusters.

    `unit` is `parent` or `store`. **Store is the quoted unit.** Seller
    calibration is a property of the store, not of the listing -- store-level
    mean deviation spans +0.25 to +1.45 across sellers
    (docs/phase1b-size-deviation-probe.md 4c) -- so two garments from one seller
    share that seller's ruler and their labelling errors are not independent
    even though the products are. Store is also the coarser partition here (68
    clusters against 92), so it gives the wider interval.

    Clusters are resampled globally rather than within stratum: 13 parents span
    more than one assigned bucket, so within-stratum resampling would break the
    cross-stratum dependence it is supposed to preserve.
    """
    rng = random.Random(seed)
    by_parent = collections.defaultdict(list)
    for row in rows:
        key = row.get(unit) or row["parent"]
        by_parent[key].append(row)
    parents = list(by_parent)

    draws = []
    for _ in range(replicates):
        sample = []
        for _ in range(len(parents)):
            sample.extend(by_parent[parents[rng.randrange(len(parents))]])
        value = statistic(sample)
        if value is not None:
            draws.append(value)
    if not draws:
        return (float("nan"), float("nan"), [])
    draws.sort()
    low = draws[int(0.025 * len(draws))]
    high = draws[min(len(draws) - 1, int(0.975 * len(draws)))]
    return (low, high, draws)


def precision(rows: list[dict], bucket: str) -> float | None:
    """Precision of one bucket, `unclear` excluded from the denominator."""
    subset = [r for r in rows if r["assigned"] == bucket and r["human"] != "unclear"]
    if not subset:
        return None
    return sum(1 for r in subset if r["human"] == bucket) / len(subset)


def lambda_of(rows: list[dict], prevalence=None, none_as_zero: bool = True):
    lam, _residual, _a = lambda_forward(confusion(rows), prevalence, none_as_zero)
    return lam


# ---------------------------------------------------------------------------
# Report
# ---------------------------------------------------------------------------

def _matrix_table(rows: list[dict], title: str) -> None:
    matrix = confusion(rows)
    columns = list(BUCKETS) + ["none", "unclear"]
    print(f"\n{title}   (n={len(rows)})")
    print(f"{'assigned \\\\ human':<20}" + "".join(f"{c:>14}" for c in columns) + f"{'total':>8}")
    for k in BUCKETS:
        cells = "".join(f"{matrix.get((k, c), 0):>14}" for c in columns)
        total = sum(matrix.get((k, c), 0) for c in columns)
        print(f"{k:<20}{cells}{total:>8}")
    totals = "".join(f"{sum(matrix.get((k, c), 0) for k in BUCKETS):>14}" for c in columns)
    print(f"{'total':<20}{totals}{len(rows):>8}")


ADJUSTMENT_FAMILY = {
    r"\bsize\s+up\b", r"\bsizing\s+up\b", r"\bsized\s+up\b",
    r"\bgo\s+(?:a\s+|one\s+|two\s+)?sizes?\s+up\b",
    r"\bgo\s+up\s+(?:a\s+|one\s+|two\s+)?sizes?\b", r"\bnext\s+size\s+up\b",
    r"\b(?:order|buy|get|purchase|choose|pick)\s+(?:a\s+|one\s+|two\s+|the\s+next\s+)?sizes?\s+(?:up|larger|bigger)\b",
    r"\b(?:order|buy|get|purchase|choose|pick)\s+(?:a\s+|one\s+|two\s+)?(?:larger|bigger)\s+sizes?\b",
    r"\bneed(?:ed)?\s+(?:a\s+)?(?:larger|bigger)\s+size\b",
    r"\bwish\s+(?:i|I)?\s*(?:had\s+)?order(?:ed)?\s+(?:a\s+)?(?:larger|bigger)\b",
    r"\bsize\s+down\b", r"\bsizing\s+down\b", r"\bsized\s+down\b",
    r"\bgo\s+(?:a\s+|one\s+|two\s+)?sizes?\s+down\b",
    r"\bgo\s+down\s+(?:a\s+|one\s+|two\s+)?sizes?\b", r"\bnext\s+size\s+down\b",
    r"\b(?:order|buy|get|purchase|choose|pick)\s+(?:a\s+|one\s+|two\s+|the\s+next\s+)?sizes?\s+(?:down|smaller)\b",
    r"\b(?:order|buy|get|purchase|choose|pick)\s+(?:a\s+|one\s+|two\s+)?smaller\s+sizes?\b",
    r"\bneed(?:ed)?\s+(?:a\s+)?smaller\s+size\b",
    r"\bwish\s+(?:i|I)?\s*(?:had\s+)?order(?:ed)?\s+(?:a\s+)?smaller\b",
}


def _drop_adjustment_family(rows: list[dict]) -> list[dict]:
    """Rows whose assignment did NOT come from an adjustment-advice pattern."""
    sys.path.insert(0, str(ROOT))
    from amazon_fit_probe import label_fit

    key = {r["review_id_hash"]: r for r in csv.DictReader(KEY.open(encoding="utf-8"))}
    by_parent_assigned = collections.defaultdict(list)
    for entry in key.values():
        by_parent_assigned[(entry["parent_asin"], entry["assigned_bucket"])].append(entry)

    kept = []
    for row in rows:
        entries = by_parent_assigned.get((row["parent"], row["assigned"]), [])
        fired = None
        for entry in entries:
            _b, hits = label_fit(entry["review_title"], entry["review_text"])
            if row["assigned"] in hits:
                fired = hits[row["assigned"]]
                break
        if fired not in ADJUSTMENT_FAMILY:
            kept.append(row)
    return kept


def main() -> int:
    rows = load_rows()
    line = "=" * 92
    print(line)
    print("ATTENUATION OF tau -- misclassification in fit_score")
    print(line)
    print(f"source: {LABELLED.name}   labelled rows: {len(rows)}")
    parents = {r["parent"] for r in rows}
    print(f"clusters: {len(parents)} parent_asin, largest carries "
          f"{max(collections.Counter(r['parent'] for r in rows).values())} rows")

    print("\n" + line)
    print("1. CONFUSION MATRICES -- raw counts")
    print(line)
    _matrix_table(rows, "ALL LABELLED ROWS")
    for gender in ("men", "women"):
        _matrix_table([r for r in rows if r["gender"] == gender],
                      f"GENDER = {gender.upper()}")
    _matrix_table(_drop_adjustment_family(rows),
                  "ADJUSTMENT-ADVICE FAMILY EXCLUDED")

    print("\n" + line)
    print("2. ATTENUATION FACTOR")
    print(line)
    subsets = [("pooled", rows),
               ("men", [r for r in rows if r["gender"] == "men"]),
               ("women", [r for r in rows if r["gender"] == "women"]),
               ("adjustment family excluded", _drop_adjustment_family(rows))]

    print(f"{'subset':<28}{'lambda_fwd':>12}{'95% CI (cluster boot)':>26}"
          f"{'lambda_rev':>12}{'residual':>11}")
    results = {}
    for name, subset in subsets:
        lam, residual, _a = lambda_forward(confusion(subset))
        rev = lambda_reverse(confusion(subset))
        if lam is None:
            print(f"{name:<28}{'--':>12}")
            continue
        low, high, _ = cluster_bootstrap(subset, lambda_of)
        results[name] = lam
        print(f"{name:<28}{lam:>12.3f}   [{low:>6.3f}, {high:>6.3f}]"
              f"{rev:>12.3f}{residual:>11.3f}")

    print("\n  lambda_fwd is operative. lambda_rev is the brief's formula, shown")
    print("  for comparison -- see the module docstring for why they differ.")
    print("  residual = a_0 - (a_small + a_large)/2; large values mean the")
    print("  single-number summary is hiding structure.")

    print("\n" + line)
    print("3. OPERATIVE MDE")
    print(line)
    for design_label, mde_design in (("0.219 (mega-listings excluded)", 0.219),
                                     ("0.568 (mega-listings kept)", 0.568)):
        print(f"\n  MDE_design = {design_label}")
        for name in ("pooled", "men", "women"):
            if name not in results:
                continue
            print(f"    {name:<26}MDE_operative = "
                  f"{mde_operative(mde_design, results[name]):.3f} SD")

    print("\n" + line)
    print("4. PRECISION -- cluster-bootstrap intervals, not Wilson")
    print(line)
    from src.analysis.score_precision import wilson

    print(f"{'bucket':<16}{'precision':>11}{'cluster bootstrap':>24}{'Wilson (iid)':>22}")
    for bucket in BUCKETS:
        point = precision(rows, bucket)
        low, high, _ = cluster_bootstrap(rows, lambda r: precision(r, bucket))
        subset = [r for r in rows if r["assigned"] == bucket and r["human"] != "unclear"]
        wlo, whi = wilson(sum(1 for r in subset if r["human"] == bucket), len(subset))
        print(f"{bucket:<16}{point:>10.1%}   [{low:>6.1%}, {high:>6.1%}]"
              f"   [{wlo:>6.1%}, {whi:>6.1%}]")
    print("\n  Four products carry more than 5% of rows each and the largest carries")
    print("  10.1%, so the iid assumption behind Wilson does not hold and the")
    print("  cluster bootstrap is the appropriate interval.")
    print("\n  BUT the percentile bootstrap is anti-conservative for proportions")
    print("  near 1 -- it cannot return an upper bound above the largest resample,")
    print("  and true_to_size sits at 59/60. Where the two disagree at a boundary,")
    print("  take the WIDER bound. Neither interval alone is right at both ends.")
    return 0


if __name__ == "__main__":
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.exit(main())
